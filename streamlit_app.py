"""
GTM Auto-Auditor - Streamlit Application
=========================================

A two-tab tool for auditing Google Tag Manager container exports.

Tab 1 — Analytics Inventory
    Upload a GTM JSON export and get a filterable table of GA4 / UA tags
    with trigger details. CSV download included.

Tab 2 — Phase 1 Audit
    Upload a GTM JSON export and run the full Phase 1 container audit.
    Produces severity-tiered findings (Critical / High / Medium / Hygiene),
    an in-browser summary, and a downloadable Excel workbook matching the
    format used by the team's manual audit process.

Dependencies
    pip install streamlit pandas openpyxl

Running locally
    streamlit run streamlit_app.py
"""

from __future__ import annotations

import hashlib
import io
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPERS  (used by both tabs)
# ══════════════════════════════════════════════════════════════════════════════

def parse_parameters(param_list: List[Dict[str, Any]]) -> Dict[str, Any]:
    param_dict: Dict[str, Any] = {}
    for param in param_list or []:
        key = param.get("key")
        if "list" in param:
            param_dict[key] = json.dumps(param.get("list"), ensure_ascii=False)
        elif "map" in param:
            param_dict[key] = json.dumps(param.get("map"), ensure_ascii=False)
        else:
            param_dict[key] = param.get("value")
    return param_dict


def extract_value(parameter: Dict[str, Any]) -> Any:
    if "value" in parameter:
        return parameter.get("value")
    if "list" in parameter:
        return [extract_value(item) for item in parameter.get("list", [])]
    if "map" in parameter:
        return {
            item.get("key"): extract_value(item)
            for item in parameter.get("map", [])
            if item.get("key")
        }
    return None


def extract_variables_from_value(value: Any) -> Set[str]:
    variables: Set[str] = set()
    if isinstance(value, str):
        variables.update(re.findall(r"\{\{([^}]+)\}\}", value))
    elif isinstance(value, list):
        for item in value:
            variables.update(extract_variables_from_value(item))
    elif isinstance(value, dict):
        for item in value.values():
            variables.update(extract_variables_from_value(item))
    return variables


def describe_trigger_type(trigger_type: Optional[str]) -> str:
    labels = {
        "PAGEVIEW": "Page View", "DOM_READY": "DOM Ready",
        "WINDOW_LOADED": "Window Loaded", "CLICK": "Click",
        "LINK_CLICK": "Link Click", "JUST_LINKS": "Link Click",
        "FORM_SUBMISSION": "Form Submission", "TIMER": "Timer",
        "SCROLL_DEPTH": "Scroll Depth", "ELEMENT_VISIBILITY": "Element Visibility",
        "CUSTOM_EVENT": "Custom Event", "YOUTUBE_VIDEO": "YouTube Video",
        "HISTORY_CHANGE": "History Change", "TRIGGER_GROUP": "Trigger Group",
        "AMP_CLICK": "AMP Click",
    }
    if not trigger_type:
        return ""
    return labels.get(trigger_type, trigger_type.replace("_", " ").title())


def describe_tag_type(tag_type: Optional[str]) -> str:
    labels = {
        "gaawe": "GA4 Event", "googtag": "GA4 Configuration",
        "ua": "Universal Analytics", "html": "Custom HTML",
        "img": "Custom Image", "gclidw": "Conversion Linker",
        "awct": "Google Ads Conversion Tracking", "sp": "Custom Template",
        "floodlight": "Floodlight",
    }
    if not tag_type:
        return ""
    return labels.get(tag_type, tag_type.replace("_", " ").title())


def describe_filter(filter_obj: Dict[str, Any]) -> str:
    filter_type = filter_obj.get("type", "")
    values = parse_parameters(filter_obj.get("parameter", []))
    arg0 = values.get("arg0", "")
    arg1 = values.get("arg1", "")
    ignore_case = str(values.get("ignore_case", "")).lower() == "true"
    operators = {
        "equals": "=", "contains": "contains", "matchRegex": "matches regex",
        "startsWith": "starts with", "endsWith": "ends with",
        "greater": ">", "less": "<",
    }
    if filter_type == "hasOwnProperty":
        return f"{arg0} exists"
    if filter_type == "cssSelector":
        return f"{arg0} matches selector {arg1}"
    operator = operators.get(filter_type, filter_type or "condition")
    description = " ".join(str(part) for part in [arg0, operator, arg1] if part)
    if ignore_case and description:
        description += " (ignore case)"
    return description


def extract_trigger_metadata(trigger: Dict[str, Any]) -> Dict[str, str]:
    custom_event_filter = parse_parameters(trigger.get("customEventFilter", []))
    filter_descriptions  = [describe_filter(f) for f in trigger.get("filter", [])]
    auto_event_descriptions = [describe_filter(f) for f in trigger.get("autoEventFilter", [])]
    trigger_conditions = " AND ".join(
        part for part in filter_descriptions + auto_event_descriptions if part
    )
    if custom_event_filter.get("arg0") or custom_event_filter.get("arg1"):
        event_match = " ".join(
            str(p) for p in [
                custom_event_filter.get("arg0"), "matches", custom_event_filter.get("arg1"),
            ] if p
        )
        trigger_conditions = " AND ".join(p for p in [event_match, trigger_conditions] if p)

    variables: Set[str] = set()
    for key in ("filter", "autoEventFilter", "customEventFilter", "parameter"):
        for item in trigger.get(key, []):
            variables.update(extract_variables_from_value(extract_value(item)))

    trigger_name = trigger.get("name", "")
    trigger_type = describe_trigger_type(trigger.get("type"))
    normalized_name = trigger_name.lower()
    all_pages = (
        "Yes"
        if "all pages" in normalized_name
        or (trigger.get("type") == "PAGEVIEW" and not trigger_conditions)
        else "No"
    )
    area_of_site = ""
    conditions_lower = trigger_conditions.lower()
    if "page path" in conditions_lower or "page url" in conditions_lower:
        area_of_site = trigger_conditions
    elif all_pages == "Yes":
        area_of_site = "Entire site"

    return {
        "Status (Live/Paused)": "Paused" if trigger.get("paused") else "Live",
        "Trigger Name": trigger_name,
        "Trigger Type": trigger_type,
        "Trigger Conditions": trigger_conditions,
        "Variables Used": ", ".join(sorted(variables)),
        "All Pages?": all_pages,
        "Area of Site": area_of_site,
    }


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — ANALYTICS INVENTORY  (original logic, unchanged)
# ══════════════════════════════════════════════════════════════════════════════

def parse_gtm_container(data: Dict[str, Any], analytics_only: bool = True) -> pd.DataFrame:
    if "containerVersion" in data:
        container = data["containerVersion"]
    else:
        container = data
    tags = container.get("tag", [])
    triggers = container.get("trigger", [])
    trigger_map: Dict[str, Dict[str, Any]] = {t.get("triggerId"): t for t in triggers}
    rows: List[Dict[str, Any]] = []
    analytics_types = {"gaawe": "GA4 Event", "googtag": "GA4 Configuration", "ua": "Universal Analytics"}

    for tag in tags:
        tag_type = tag.get("type")
        if analytics_only and tag_type not in analytics_types:
            continue
        tag_name = tag.get("name", "")
        type_label = analytics_types.get(tag_type, describe_tag_type(tag_type))
        firing_ids = tag.get("firingTriggerId", [])
        param_dict = parse_parameters(tag.get("parameter", []))
        event_name     = param_dict.get("eventName")
        event_category = param_dict.get("eventCategory") or param_dict.get("category")
        event_action   = param_dict.get("eventAction")   or param_dict.get("action")
        event_label    = param_dict.get("eventLabel")    or param_dict.get("label")
        exclude_keys = {"eventName","eventCategory","category","eventAction","action","eventLabel","label"}
        other_params = {k: v for k, v in param_dict.items() if k not in exclude_keys and v not in (None, "")}
        other_params_str = "; ".join(f"{k}={v}" for k, v in other_params.items())
        variables_used = set(extract_variables_from_value(other_params))
        event_or_action = " | ".join(
            str(v) for v in [event_name, event_category, event_action, event_label] if v not in (None, "")
        )
        if not firing_ids:
            firing_ids = [None]
        for trigger_id in firing_ids:
            trigger = trigger_map.get(trigger_id, {})
            trigger_metadata = (
                extract_trigger_metadata(trigger) if trigger else {
                    "Status (Live/Paused)": "Live",
                    "Trigger Name": f"ID:{trigger_id}" if trigger_id else "",
                    "Trigger Type": "", "Trigger Conditions": "",
                    "Variables Used": "", "All Pages?": "No", "Area of Site": "",
                }
            )
            combined_variables = sorted({
                *variables_used,
                *(set(trigger_metadata["Variables Used"].split(", ")) if trigger_metadata["Variables Used"] else set()),
            })
            rows.append({
                "Tag Name": tag_name, "Tag Type": type_label,
                "Status (Live/Paused)": "Paused" if tag.get("paused") else trigger_metadata["Status (Live/Paused)"],
                "Trigger Name": trigger_metadata["Trigger Name"],
                "Trigger Type": trigger_metadata["Trigger Type"],
                "Trigger Conditions": trigger_metadata["Trigger Conditions"],
                "Variables Used": ", ".join(v for v in combined_variables if v),
                "Event Name / Action": event_or_action,
                "All Pages?": trigger_metadata["All Pages?"],
                "Area of Site": trigger_metadata["Area of Site"],
                "Event Name": event_name, "Event Category": event_category,
                "Event Action": event_action, "Event Label": event_label,
                "Parameters": other_params_str,
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    preferred_columns = [
        "Status (Live/Paused)", "Trigger Name", "Trigger Type", "Trigger Conditions",
        "Variables Used", "Event Name / Action", "All Pages?", "Area of Site",
        "Tag Name", "Tag Type", "Event Name", "Event Category",
        "Event Action", "Event Label", "Parameters",
    ]
    return df[[c for c in preferred_columns if c in df.columns]]


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — PHASE 1 AUDIT ENGINE  (ported from gtm_auditor.py)
# ══════════════════════════════════════════════════════════════════════════════

ALL_PAGES_TRIGGERS = {"2147479553", "2147479573", "2147479574"}

VENDOR_KEYWORDS = [
    ("facebook",     "Meta/Facebook"),  ("meta pixel",   "Meta/Facebook"),
    ("linkedin",     "LinkedIn"),       ("twitter",      "Twitter/X"),
    ("videoamp",     "VideoAmp"),       ("adtheorent",   "AdTheorent"),
    ("equativ",      "Equativ"),        ("nexxen",       "Nexxen"),
    ("reddit",       "Reddit"),         ("onetrust",     "OneTrust / Consent"),
    ("qualtrics",    "Qualtrics"),      ("resonate",     "Resonate"),
    ("silkroad",     "SilkRoad"),       ("envestnet",    "Envestnet"),
    ("bing",         "Microsoft Bing Ads"), ("hotjar",   "Hotjar"),
    ("lucky orange", "Lucky Orange"),   ("pinterest",    "Pinterest"),
    ("tiktok",       "TikTok"),         ("snapchat",     "Snapchat"),
    ("criteo",       "Criteo"),         ("taboola",      "Taboola"),
    ("outbrain",     "Outbrain"),       ("hubspot",      "HubSpot"),
    ("marketo",      "Marketo"),        ("pardot",       "Pardot"),
    ("salesforce",   "Salesforce"),     ("intercom",     "Intercom"),
    ("drift",        "Drift"),          ("zendesk",      "Zendesk"),
    ("segment",      "Segment"),        ("heap",         "Heap"),
    ("mixpanel",     "Mixpanel"),       ("amplitude",    "Amplitude"),
    ("fullstory",    "FullStory"),      ("clarity",      "Microsoft Clarity"),
    ("cookie",       "Consent / Cookie"), ("consent",   "Consent / Cookie"),
    ("cookiebot",    "Cookiebot"),      ("trustarc",     "TrustArc"),
    ("usercentrics", "Usercentrics"),   ("didomi",       "Didomi"),
    ("mc personal",  "Salesforce MC"), ("braze",        "Braze"),
]

TEMPLATE_VENDOR_MAP = {
    "cvt_PBGZL": "Reddit", "bzi": "LinkedIn Insight Tag",
    "baut": "Microsoft Bing Ads", "twitter_website_tag": "Twitter/X",
    "pntr": "Pinterest", "hjtc": "Hotjar", "lcl": "Lucky Orange",
}


def _get_param(tag: Dict[str, Any], key: str, default: str = "") -> str:
    for p in tag.get("parameter", []):
        if p.get("key") == key:
            return p.get("value", default)
    return default


def _classify_vendor(tag: Dict[str, Any]) -> str:
    name = tag.get("name", "").lower()
    typ  = tag.get("type", "")
    if typ == "ua":     return "Universal Analytics"
    if typ == "flc":    return "Floodlight (DV360)"
    if typ == "fls":    return "Floodlight (DV360)"
    if typ == "gclidw": return "Google Ads Conversion Linker"
    if typ == "awct":   return "Google Ads Conversion"
    if typ == "awud":   return "Google Ads User Data"
    if typ == "awdc":   return "Google Ads Remarketing"
    if typ == "opt":    return "Google Optimize"
    if typ == "googtag":
        tag_id_val = _get_param(tag, "tagId")
        if tag_id_val.startswith("AW-"): return "Google Ads"
        if tag_id_val.startswith("DC-"): return "Floodlight (DV360)"
        if tag_id_val.startswith("G-"):
            if any(kw in name for kw in ("dev", "test", "staging")): return "GA4 (dev/test label)"
            return "GA4"
        return "GA4"
    if typ == "gaawe": return "GA4"
    for prefix, vendor in TEMPLATE_VENDOR_MAP.items():
        if typ == prefix or typ.startswith(f"cvt_{prefix}"):
            return vendor
    if typ.startswith("cvt_"):
        for kw, vendor in VENDOR_KEYWORDS:
            if kw in name:
                return vendor
        return "Custom Template Tag"
    for kw, vendor in VENDOR_KEYWORDS:
        if kw in name:
            return vendor
    if typ == "html":
        return "Custom HTML"
    return typ if typ else "Other"


def _load_scope(tag: Dict[str, Any], trig_names: Dict[str, str]) -> str:
    fires = tag.get("firingTriggerId", [])
    if not fires:
        return "Orphaned (no trigger)"
    if any(tid in ALL_PAGES_TRIGGERS for tid in fires):
        return "All Pages"
    names_lower = [trig_names.get(tid, "").lower() for tid in fires]
    if any(kw in n for n in names_lower for kw in ("all page", "all-page", "pageview", "page view")):
        return "Broad page group"
    if any(kw in n for n in names_lower for kw in ("submit", "complete", "done", "confirm", "thank", "conversion", "purchase")):
        return "Conversion-only"
    if any(kw in n for n in names_lower for kw in ("click", "scroll", "video", "button", "link")):
        return "Interaction-based"
    return "Conditional / specific"


def _build_trigger_map(triggers: List[Dict[str, Any]]):
    names = {tr["triggerId"]: tr["name"] for tr in triggers}
    names.update({
        "2147479553": "All Pages (gtm.load)",
        "2147479573": "All Pages (gtm.js)",
        "2147479574": "All Pages (gtm.dom)",
    })
    return names


def _get_used_trigger_ids(tags: List[Dict[str, Any]]) -> Set[str]:
    used: Set[str] = set()
    for t in tags:
        for tid in t.get("firingTriggerId", []): used.add(tid)
        for tid in t.get("blockingTriggerId", []): used.add(tid)
    return used


def _load_container_dict(data: Dict[str, Any]) -> Dict[str, Any]:
    cv = data.get("containerVersion", data)
    meta = cv.get("container", {})
    return {
        "container_id":   meta.get("publicId", cv.get("containerId", "Unknown")),
        "container_name": meta.get("name", "Unknown"),
        "export_time":    data.get("exportTime", "Unknown"),
        "tags":           cv.get("tag", []),
        "triggers":       cv.get("trigger", []),
        "variables":      cv.get("variable", []),
        "custom_templates": cv.get("customTemplate", []),
    }


def generate_findings(container: Dict[str, Any]) -> List[Dict[str, Any]]:
    tags     = container["tags"]
    triggers = container["triggers"]
    trig_names = _build_trigger_map(triggers)
    used_tids  = _get_used_trigger_ids(tags)

    live_tags   = [t for t in tags if not t.get("paused")]
    paused_tags = [t for t in tags if t.get("paused")]
    findings: List[Dict[str, Any]] = []
    sev_counter: Dict[str, int] = {"Critical": 0, "High": 0, "Medium": 0, "Hygiene": 0}

    def add(severity, category, title, affected, why, action, owner="Analytics"):
        sev_counter[severity] = sev_counter.get(severity, 0) + 1
        prefix = {"Critical": "C", "High": "H", "Medium": "M", "Hygiene": "Hy"}[severity]
        findings.append({
            "id": f"{prefix}-{sev_counter[severity]}",
            "severity": severity, "category": category,
            "title": title, "affected": affected,
            "why": why, "action": action, "owner": owner,
        })

    # ── GA4 CONFIG ─────────────────────────────────────────────────────────
    ga4_configs = [t for t in live_tags if t["type"] == "googtag"
                   and _get_param(t, "tagId").startswith("G-")]
    all_pages_ga4 = [t for t in ga4_configs
                     if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))]

    if not ga4_configs:
        add("Critical", "Measurement",
            "No GA4 configuration tag found",
            "No live googtag with a G-XXXXXXXX measurement ID",
            "Without a GA4 config tag, all GA4 event tags are broken — data cannot be sent to any GA4 property.",
            "Add a GA4 configuration tag with the correct production measurement ID firing on All Pages.")
    elif not all_pages_ga4:
        add("High", "Measurement",
            f"{len(ga4_configs)} GA4 config tag(s) found but none fire on all pages",
            "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in ga4_configs),
            "A GA4 config tag that doesn't fire on all pages will cause GA4 event tags to fail silently on uncovered pages.",
            "Ensure the primary GA4 config tag fires on All Pages.")
    else:
        dev_configs = [t for t in all_pages_ga4
                       if any(kw in t["name"].lower() for kw in ("dev", "test", "staging", "local"))]
        if dev_configs and len(dev_configs) == len(all_pages_ga4):
            ids = ", ".join(_get_param(t, "tagId") for t in dev_configs)
            add("Critical", "Measurement",
                "GA4 config tag is labeled dev/test — production measurement unconfirmed",
                "; ".join(f"Tag {t['tagId']}: {t['name']} ({_get_param(t, 'tagId')})" for t in dev_configs),
                f"The only all-pages GA4 config tag(s) carry a dev/test label ({ids}). All 26+ GA4 event tags route data to this property. If it is a development property, production analytics is dark.",
                f"Open GA4 Admin and confirm whether {ids} is the production property. If dev: add a production config tag immediately. If prod: rename to remove the dev label and add a GTM note.")

        unique_ids = set(_get_param(t, "tagId") for t in all_pages_ga4)
        if len(unique_ids) > 1:
            details = "; ".join(f"Tag {t['tagId']}: {t['name']} → {_get_param(t, 'tagId')}" for t in all_pages_ga4)
            add("High", "Measurement",
                f"Multiple GA4 config tags on all pages with different measurement IDs ({len(unique_ids)} IDs)",
                details,
                "Multiple GA4 configs with different IDs can split pageviews and events across properties, causing double-counting or data gaps.",
                "Confirm the primary production measurement ID. Scope secondary configs to specific page sets and add GTM notes documenting the architecture.")

    # ── UNIVERSAL ANALYTICS ────────────────────────────────────────────────
    ua_live   = [t for t in live_tags   if t["type"] == "ua"]
    ua_paused = [t for t in paused_tags if t["type"] == "ua"]

    if ua_live:
        ua_ap = [t for t in ua_live if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))]
        details = "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in ua_live[:8])
        if len(ua_live) > 8:
            details += f" … and {len(ua_live) - 8} more"
        add("Critical" if ua_ap else "High", "Legacy",
            f"{len(ua_live)} live Universal Analytics tag(s) still firing — UA sunset July 2023",
            details,
            f"Google sunset Universal Analytics in July 2023. These {len(ua_live)} live tags send data to a property that no longer processes it — pure page-load overhead with no measurement return."
            + (f" Tag(s) {', '.join(t['tagId'] for t in ua_ap)} fire on every page." if ua_ap else ""),
            f"Pause and remove all {len(ua_live)} live UA tags. Verify all measurement objectives have been migrated to GA4 first.")

    if ua_paused:
        details = "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in ua_paused)
        add("Hygiene", "Cleanup",
            f"{len(ua_paused)} paused Universal Analytics tag(s) still present",
            details,
            "Paused legacy tags create confusion. They should be deleted, not merely paused.",
            "Remove all paused UA tags once migration to GA4 is confirmed complete.")

    # ── DUPLICATE ALL-PAGES BASE TAGS BY VENDOR ────────────────────────────
    ap_live = [t for t in live_tags if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))]
    skip_vendors = {
        "GA4", "GA4 (dev/test label)", "Google Ads", "Google Ads Conversion Linker",
        "Floodlight (DV360)", "Google Ads Remarketing", "Google Optimize",
        "Microsoft Bing Ads", "GA Scrolling Pages", "Universal Analytics",
    }
    vendor_ap: Dict[str, list] = defaultdict(list)
    for t in ap_live:
        v = _classify_vendor(t)
        if v not in skip_vendors:
            vendor_ap[v].append(t)

    for vendor, vtags in vendor_ap.items():
        if len(vtags) >= 2:
            details = "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in vtags)
            add("High", "Performance",
                f"Duplicate all-pages base tags for {vendor} ({len(vtags)} tags loading on every page)",
                details,
                f"{vendor} has {len(vtags)} separate tags all firing sitewide. This multiplies vendor network overhead on every page load and may produce duplicate signals.",
                f"Confirm with the {vendor} team whether multiple all-pages tags are intentional. Consolidate to one base tag where possible.",
                "Media")

    # ── CUSTOM HTML VOLUME ─────────────────────────────────────────────────
    html_live = [t for t in live_tags if t["type"] == "html"]
    html_ap   = [t for t in html_live if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))]

    if len(html_live) >= 10:
        pct = round(len(html_live) / max(len(live_tags), 1) * 100)
        add("High", "Governance",
            f"{len(html_live)} live Custom HTML tags ({pct}% of all live tags) — {len(html_ap)} on all pages",
            f"{len(html_ap)} all-pages Custom HTML tags; {len(html_live)} live Custom HTML total",
            "Custom HTML tags bypass GTM's consent and configuration framework, inject arbitrary external scripts, and are harder to audit and maintain.",
            "Audit each Custom HTML tag for template-based equivalents. Prioritize migrating all-pages Custom HTML. Add GTM notes documenting purpose and owner for all remaining Custom HTML.",
            "Engineering")
    elif len(html_live) >= 3:
        add("Medium", "Governance",
            f"{len(html_live)} live Custom HTML tags — review for template alternatives",
            f"{len(html_ap)} on all pages; {len(html_live)} total",
            "Custom HTML tags require higher governance attention than template-based equivalents.",
            "Review each Custom HTML tag and migrate to a template where one exists.",
            "Engineering")

    # ── ORPHANED LIVE TAGS ─────────────────────────────────────────────────
    orphaned = [t for t in live_tags if not t.get("firingTriggerId")]
    if orphaned:
        details = "; ".join(f"Tag {t['tagId']}: {t['name']} ({t['type']})" for t in orphaned)
        add("High" if len(orphaned) > 3 else "Medium", "Cleanup",
            f"{len(orphaned)} live tag(s) with no firing trigger — will never execute",
            details,
            "Tags with no firing trigger are dead configuration. They contribute to container clutter.",
            "Remove orphaned tags or assign appropriate firing triggers if they should be active.")

    # ── CONSENT / ONETRUST ─────────────────────────────────────────────────
    consent_tags = [t for t in tags if any(
        kw in t.get("name", "").lower()
        for kw in ("onetrust", "cookiebot", "consent", "cookie banner", "trustarc", "usercentrics", "didomi")
    )]
    consent_paused = [t for t in consent_tags if t.get("paused")]
    consent_live   = [t for t in consent_tags if not t.get("paused")]
    no_blocking    = [t for t in live_tags if not t.get("blockingTriggerId")]

    if consent_paused and not consent_live:
        add("Medium", "Governance",
            f"Consent management tag(s) are paused — consent enforcement may be inactive",
            "; ".join(f"Tag {t['tagId']}: {t['name']} (PAUSED)" for t in consent_paused),
            f"The consent management tag is paused. {len(no_blocking)} live tags have no blocking triggers — vendor pixels may fire before consent.",
            "Confirm with engineering whether consent is managed outside GTM. If GTM-based, reactivate the consent tag and add blocking triggers to all vendor pixel tags.",
            "Privacy / Engineering")
    elif len(no_blocking) > 20 and not consent_tags:
        add("Medium", "Governance",
            f"No consent management tag detected — {len(no_blocking)} live tags have no consent blocking",
            f"{len(no_blocking)} live tags lack blocking triggers",
            "Without a consent management platform or blocking triggers, vendor pixels may fire before user consent.",
            "Implement a CMP (OneTrust, Cookiebot, etc.) and add blocking triggers to all vendor pixel tags.",
            "Privacy / Engineering")

    # ── HIGH-DENSITY CONVERSION TRIGGER ───────────────────────────────────
    trig_tag_counts: Counter = Counter()
    trig_tag_names_map: Dict[str, list] = defaultdict(list)
    for t in live_tags:
        for tid in t.get("firingTriggerId", []):
            trig_tag_counts[tid] += 1
            trig_tag_names_map[tid].append(t["name"])

    for tid, count in trig_tag_counts.items():
        if count >= 8 and tid not in ALL_PAGES_TRIGGERS:
            tname = trig_names.get(tid, f"Trigger {tid}")
            vendors = [_classify_vendor(t) for t in live_tags if tid in t.get("firingTriggerId", [])]
            dupes = {v: c for v, c in Counter(vendors).items() if c >= 2}
            dupe_note = f" Potential duplicate signals: {', '.join(f'{v} ×{c}' for v, c in dupes.items())}." if dupes else ""
            tag_list = ", ".join(trig_tag_names_map[tid][:6])
            if count > 6:
                tag_list += "…"
            add("Medium", "Measurement",
                f"{count} tags fire on '{tname}' — review for duplicate conversion signals",
                f"Trigger: {tname} (ID {tid}); Tags: {tag_list}",
                f"{count} tags fire simultaneously on this trigger.{dupe_note}",
                f"Confirm all {count} tags are intentional. Check for duplicate vendor conversion signals. Document tag purposes in GTM notes.",
                "Media / Analytics")

    # ── UNUSED TRIGGERS ────────────────────────────────────────────────────
    triggers_list = container["triggers"]
    unused = [tr for tr in triggers_list if tr["triggerId"] not in used_tids]
    unused_pct = round(len(unused) / max(len(triggers_list), 1) * 100)

    if len(unused) >= 20:
        add("High", "Cleanup",
            f"{len(unused)} unused triggers ({unused_pct}% of all triggers)",
            f"{len(unused)} triggers have no tags attached",
            f"Over {unused_pct}% of triggers are unused. This significantly inflates the container's audit surface and maintenance burden.",
            "Remove all unused triggers in a cleanup sprint. Check for any referenced in custom JS variables first.")
    elif len(unused) >= 5:
        sev = "Medium" if unused_pct >= 20 else "Hygiene"
        add(sev, "Cleanup",
            f"{len(unused)} unused triggers ({unused_pct}% of all triggers)",
            f"{len(unused)} triggers have no tags attached",
            "Unused triggers add clutter and make the container harder to navigate.",
            "Remove unused triggers in next cleanup sprint.")
    elif unused:
        add("Hygiene", "Cleanup",
            f"{len(unused)} unused trigger(s)",
            f"{len(unused)} triggers with no tags",
            "Minor trigger clutter.",
            "Remove unused triggers.")

    # ── NAMING HYGIENE ─────────────────────────────────────────────────────
    bad_pattern = re.compile(
        r'\b(test|temp|old|backup|copy|new|delete|del|todo|fix|fixme|xxx|deprecated|draft|wip)\b',
        re.IGNORECASE,
    )
    bad_names = [t for t in live_tags if bad_pattern.search(t.get("name", ""))]
    stale_container = bool(re.search(
        r'\[(new|old|test|temp|copy|bak|backup|draft)\]',
        container.get("container_name", ""), re.IGNORECASE,
    ))
    if len(bad_names) >= 5 or stale_container:
        name_note = f"Container name: '{container['container_name']}'. " if stale_container else ""
        tag_note  = f"{len(bad_names)} tags with naming issues: {', '.join(t['name'] for t in bad_names[:5])}{'…' if len(bad_names) > 5 else ''}" if bad_names else ""
        add("Hygiene", "Governance",
            "Container and/or tag naming hygiene issues",
            (name_note + tag_note).strip(),
            "Stale naming conventions (test, old, new, temp, copy) create confusion about what is active.",
            "Rename the container and any tags with stale labels. Establish a naming convention.")

    # ── HARDCODED IDs IN CUSTOM HTML ───────────────────────────────────────
    hardcoded = []
    for t in html_live:
        html_content = _get_param(t, "html")
        if re.search(r'UA-\d{4,}-\d+|G-[A-Z0-9]{8,}|AW-\d{7,}', html_content):
            hardcoded.append(t)
    if hardcoded:
        details = "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in hardcoded[:5])
        add("Medium", "Governance",
            f"{len(hardcoded)} Custom HTML tag(s) contain hardcoded tracking IDs",
            details,
            "Hardcoded IDs in Custom HTML make ID changes error-prone and increase the risk of stale configuration.",
            "Replace hardcoded IDs with GTM Constant variables so they can be managed centrally.",
            "Engineering")

    # ── MULTIPLE CONSENT TOOLS ─────────────────────────────────────────────
    consent_vendors: Set[str] = set()
    for t in consent_tags:
        n = t.get("name", "").lower()
        for kw, v in [("onetrust","OneTrust"),("cookiebot","Cookiebot"),("trustarc","TrustArc"),
                      ("usercentrics","Usercentrics"),("didomi","Didomi")]:
            if kw in n:
                consent_vendors.add(v)
    if len(consent_vendors) >= 2:
        add("High", "Governance",
            f"Multiple consent tools detected: {', '.join(sorted(consent_vendors))}",
            "; ".join(f"Tag {t['tagId']}: {t['name']}" for t in consent_tags),
            "Multiple consent management platforms produce conflicting consent signals and add unnecessary code weight.",
            f"Confirm which consent tool is authoritative and remove or disable the other(s).",
            "Privacy / Engineering")

    sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Hygiene": 3}
    return sorted(findings, key=lambda f: sev_order.get(f["severity"], 9))


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK BUILDER  (in-memory, no disk writes)
# ══════════════════════════════════════════════════════════════════════════════

NAVY, NAVY2 = "0D1B3E", "1A2744"
RED, AMB, BLUE_C, GREEN_C = "C0392B", "D4881C", "1565C0", "1A6B3C"
GRAY_C, WHITE_C = "566573", "FFFFFF"
LT_GRY, LT_RED, LT_AMB, LT_BLU, LT_GRN = "F2F3F4", "FADBD8", "FDEBD0", "D6EAF8", "D5F5E3"
SEV_BG = {"Critical": LT_RED, "High": LT_AMB, "Medium": LT_BLU, "Hygiene": LT_GRN}
SEV_FG = {"Critical": RED,    "High": AMB,     "Medium": BLUE_C,  "Hygiene": GREEN_C}


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _borders(ws, r1, r2, c1, c2):
    b = _thin_border()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title(ws, text, ncols):
    ws.row_dimensions[1].height = 26
    c = ws.cell(row=1, column=1, value=text)
    c.font      = Font(name="Arial", size=12, bold=True, color=WHITE_C)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

def _hdr(ws, row, col, val, bg=NAVY, fg=WHITE_C, bold=True, sz=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=sz, bold=bold, color=fg)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    return c

def _cell(ws, row, col, val="", bg=WHITE_C, bold=False, color="000000", sz=9, align="left", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=sz, bold=bold, color=color)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    return c


def build_audit_workbook(container: Dict[str, Any], findings: List[Dict[str, Any]]) -> bytes:
    """Build the full audit workbook and return raw bytes for st.download_button."""
    tags     = container["tags"]
    triggers = container["triggers"]
    trig_names = _build_trigger_map(triggers)
    used_tids  = _get_used_trigger_ids(tags)
    live_tags   = [t for t in tags if not t.get("paused")]
    paused_tags = [t for t in tags if t.get("paused")]
    unused_trigs = [tr for tr in triggers if tr["triggerId"] not in used_tids]
    ap_all = [t for t in tags if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))]
    sev_counts = Counter(f["severity"] for f in findings)

    wb = Workbook()

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    _title(ws1, f"GTM Audit  |  {container['container_id']}  |  {container['container_name']}  |  Phase 1 Container Audit", 4)

    meta = [
        ("Container ID",       container["container_id"]),
        ("Container Name",     container["container_name"]),
        ("Export Date",        str(container["export_time"])[:19]),
        ("Audit Generated",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Tags",         str(len(tags))),
        ("Live Tags",          str(len(live_tags))),
        ("Paused Tags",        str(len(paused_tags))),
        ("Total Triggers",     str(len(triggers))),
        ("Unused Triggers",    f"{len(unused_trigs)} ({round(len(unused_trigs)/max(len(triggers),1)*100)}%)"),
        ("Total Variables",    str(len(container["variables"]))),
        ("", ""),
        ("All-Pages Live Tags", str(len([t for t in live_tags if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId",[]))]))),
        ("Custom HTML Tags",   str(len([t for t in tags if t["type"] == "html"]))),
        ("UA Tags (live)",     str(len([t for t in live_tags if t["type"] == "ua"]))),
        ("GA4 Config Tags",    str(len([t for t in live_tags if t["type"] == "googtag" and _get_param(t, "tagId").startswith("G-")]))),
    ]
    for r, (lbl, val) in enumerate(meta, 3):
        if not lbl:
            ws1.row_dimensions[r].height = 7; continue
        ws1.row_dimensions[r].height = 15
        _cell(ws1, r, 1, lbl, bg=LT_GRY, bold=True, sz=9, wrap=False)
        c = ws1.cell(row=r, column=2, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    r = len(meta) + 5
    for col, h in enumerate(["Severity", "Findings", "Summary"], 1):
        _hdr(ws1, r, col, h, bg=NAVY2)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    sev_summaries: Dict[str, list] = {s: [] for s in ["Critical","High","Medium","Hygiene"]}
    for f in findings:
        sev_summaries[f["severity"]].append(f["title"])

    for sev in ["Critical","High","Medium","Hygiene"]:
        r += 1
        ws1.row_dimensions[r].height = 44
        count   = sev_counts.get(sev, 0)
        summary = "; ".join(sev_summaries[sev]) or "No findings"
        bg, fg  = SEV_BG[sev], SEV_FG[sev]
        _cell(ws1, r, 1, sev,        bg=bg, bold=True, color=fg, sz=9, wrap=False)
        _cell(ws1, r, 2, str(count), bg=bg, bold=True, color=fg, sz=11, align="center", wrap=False)
        c = ws1.cell(row=r, column=3, value=summary)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    _borders(ws1, 3, r, 1, 4)
    _widths(ws1, [26, 14, 50, 16])

    # ── Sheet 2: Findings Matrix ───────────────────────────────────────────
    ws2 = wb.create_sheet("Findings Matrix")
    ws2.sheet_view.showGridLines = False
    _title(ws2, f"Findings Matrix  |  {container['container_id']}  |  {len(findings)} findings", 8)
    for i, col in enumerate(["ID","Severity","Category","Finding","Affected Tags / Triggers",
                              "Why It Matters","Recommended Action","Owner"], 1):
        _hdr(ws2, 2, i, col)
    for i, f in enumerate(findings, 3):
        ws2.row_dimensions[i].height = 72
        bg, fg = SEV_BG[f["severity"]], SEV_FG[f["severity"]]
        alt = LT_GRY if i % 2 == 0 else WHITE_C
        _cell(ws2, i, 1, f["id"],       bg=bg,  bold=True, color=fg, sz=9, wrap=False)
        _cell(ws2, i, 2, f["severity"], bg=bg,  bold=True, color=fg, sz=9, wrap=False)
        _cell(ws2, i, 3, f["category"], bg=alt, sz=9, wrap=False)
        _cell(ws2, i, 4, f["title"],    bg=alt, bold=True, sz=9)
        _cell(ws2, i, 5, f["affected"], bg=alt, sz=9)
        _cell(ws2, i, 6, f["why"],      bg=alt, sz=9)
        _cell(ws2, i, 7, f["action"],   bg=alt, sz=9)
        _cell(ws2, i, 8, f["owner"],    bg=alt, sz=9, wrap=False)
    _borders(ws2, 2, len(findings) + 2, 1, 8)
    _widths(ws2, [6, 9, 12, 42, 44, 48, 50, 16])
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Tag Inventory ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Tag Inventory")
    ws3.sheet_view.showGridLines = False
    _title(ws3, f"Tag Inventory  |  {container['container_id']}  |  {len(tags)} tags", 8)
    for i, col in enumerate(["Tag ID","Tag Name","Vendor / Platform","GTM Type",
                              "Status","Load Scope","Firing Triggers","Blocking Triggers"], 1):
        _hdr(ws3, 2, i, col)
    for ri, t in enumerate(tags, 3):
        ws3.row_dimensions[ri].height = 28
        is_orphan = not t.get("firingTriggerId") and not t.get("paused")
        status = "Paused" if t.get("paused") else ("Orphaned" if is_orphan else "Live")
        fire_str  = ", ".join(trig_names.get(tid, f"#{tid}") for tid in t.get("firingTriggerId",  []))
        block_str = ", ".join(trig_names.get(tid, f"#{tid}") for tid in t.get("blockingTriggerId", []))
        bg  = LT_RED if is_orphan else (LT_AMB if t.get("paused") else (LT_GRY if ri%2==0 else WHITE_C))
        sfg = GREEN_C if status=="Live" else (AMB if status=="Paused" else RED)
        _cell(ws3, ri, 1, t["tagId"],        bg=bg, sz=9, wrap=False)
        _cell(ws3, ri, 2, t["name"],         bg=bg, sz=9, bold=(status=="Live"))
        _cell(ws3, ri, 3, _classify_vendor(t), bg=bg, sz=9, wrap=False)
        _cell(ws3, ri, 4, t["type"],         bg=bg, sz=9, wrap=False)
        _cell(ws3, ri, 5, status,            bg=bg, sz=9, bold=True, color=sfg, wrap=False)
        _cell(ws3, ri, 6, _load_scope(t, trig_names), bg=bg, sz=9, wrap=False)
        _cell(ws3, ri, 7, fire_str  or "—", bg=bg, sz=8)
        _cell(ws3, ri, 8, block_str or "—", bg=bg, sz=8)
    _borders(ws3, 2, len(tags) + 2, 1, 8)
    _widths(ws3, [8, 54, 24, 16, 10, 22, 64, 40])
    ws3.freeze_panes = "A3"

    # ── Sheet 4: All-Pages Load Review ─────────────────────────────────────
    ws4 = wb.create_sheet("All-Pages Load Review")
    ws4.sheet_view.showGridLines = False
    _title(ws4, f"All-Pages Load Review  |  {len(ap_all)} tags fire on every page", 6)
    for i, col in enumerate(["Tag ID","Tag Name","Vendor","Type","Status","Performance / Governance Flag"], 1):
        _hdr(ws4, 2, i, col)
    for ri, t in enumerate(ap_all, 3):
        ws4.row_dimensions[ri].height = 36
        vendor  = _classify_vendor(t)
        is_html = t["type"] == "html"
        is_ua   = t["type"] == "ua"
        has_con = bool(t.get("blockingTriggerId"))
        is_dev  = any(kw in t.get("name","").lower() for kw in ("dev","test","staging"))
        status  = "Paused" if t.get("paused") else "Live"
        if is_ua:
            flag, sev = "Critical — UA tag post-sunset. Remove immediately.", "Critical"
        elif is_dev:
            flag, sev = "Critical — Dev/test label on all-pages tag. Verify production intent.", "Critical"
        elif is_html and not has_con:
            flag, sev = "High — Custom HTML with no consent blocking.", "High"
        elif is_html and has_con:
            flag, sev = "Medium — Custom HTML with consent blocking. Evaluate template alternative.", "Medium"
        elif not has_con and not t.get("paused"):
            flag, sev = "Medium — No consent blocking trigger.", "Medium"
        elif t.get("paused"):
            flag, sev = "Low — Tag is paused. Remove if no longer needed.", "Low"
        else:
            flag, sev = "Review — Template-based global tag. Confirm active vendor relationship.", "Review"
        bg = LT_RED if sev=="Critical" else (LT_AMB if sev=="High" else (LT_BLU if sev=="Medium" else LT_GRY))
        fg = RED   if sev=="Critical" else (AMB  if sev=="High" else (BLUE_C if sev=="Medium" else GRAY_C))
        _cell(ws4, ri, 1, t["tagId"], bg=bg, sz=9, wrap=False)
        _cell(ws4, ri, 2, t["name"],  bg=bg, sz=9, bold=True)
        _cell(ws4, ri, 3, vendor,     bg=bg, sz=9, wrap=False)
        _cell(ws4, ri, 4, t["type"],  bg=bg, sz=9, wrap=False)
        _cell(ws4, ri, 5, status,     bg=bg, sz=9, bold=True, color=GREEN_C if not t.get("paused") else GRAY_C, wrap=False)
        _cell(ws4, ri, 6, flag,       bg=bg, bold=True, color=fg, sz=9)
    _borders(ws4, 2, len(ap_all) + 2, 1, 6)
    _widths(ws4, [8, 54, 24, 14, 10, 72])
    ws4.freeze_panes = "A3"

    # ── Sheet 5: Vendor Summary ────────────────────────────────────────────
    ws5 = wb.create_sheet("Vendor Summary")
    ws5.sheet_view.showGridLines = False
    _title(ws5, f"Vendor Summary  |  {container['container_id']}", 6)
    for i, col in enumerate(["Vendor / Platform","Live Tags","All-Pages","Custom HTML?",
                              "Consent Blocked?","Notes"], 1):
        _hdr(ws5, 2, i, col)
    vd: Dict[str, Any] = defaultdict(lambda: {"live":0,"ap":0,"html":0,"blocked":0,"unblocked":0})
    for t in tags:
        v = _classify_vendor(t)
        if not t.get("paused"):
            vd[v]["live"] += 1
            if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", [])): vd[v]["ap"] += 1
            if t["type"] == "html": vd[v]["html"] += 1
            if t.get("blockingTriggerId"): vd[v]["blocked"] += 1
            else: vd[v]["unblocked"] += 1
    for ri, (vendor, d) in enumerate(sorted(vd.items(), key=lambda x: -x[1]["live"]), 3):
        ws5.row_dimensions[ri].height = 36
        cons = "Mostly yes" if d["blocked"] > d["unblocked"] else ("No" if d["blocked"]==0 else "Partial")
        notes = []
        if vendor == "Universal Analytics" and d["live"] > 0:
            notes.append(f"CRITICAL: {d['live']} live UA tags post-sunset. Remove immediately.")
        if d["ap"] >= 2:
            notes.append(f"HIGH: {d['ap']} all-pages tags — review for duplicates.")
        if d["html"] >= 5:
            notes.append(f"High Custom HTML volume ({d['html']} tags).")
        if d["ap"] > 0 and d["blocked"] == 0 and vendor not in ("GA4","Universal Analytics","Google Ads","GA4 (dev/test label)"):
            notes.append("No consent blocking on any tags.")
        if not notes:
            notes.append("Review active vendor relationship and tag governance.")
        note = " ".join(notes)
        bg = LT_RED if "CRITICAL" in note else (LT_AMB if "HIGH" in note else (LT_GRY if ri%2==0 else WHITE_C))
        _cell(ws5, ri, 1, vendor,               bg=bg, bold=True, sz=9, wrap=False)
        _cell(ws5, ri, 2, str(d["live"]),        bg=bg, sz=9, align="center", wrap=False)
        _cell(ws5, ri, 3, str(d["ap"]) if d["ap"] else "—", bg=bg, sz=9, align="center", wrap=False)
        _cell(ws5, ri, 4, "Yes" if d["html"] else "No", bg=bg, sz=9, align="center", wrap=False)
        _cell(ws5, ri, 5, cons,                 bg=bg, sz=9, wrap=False)
        _cell(ws5, ri, 6, note,                 bg=bg, sz=9)
    _borders(ws5, 2, len(vd) + 2, 1, 6)
    _widths(ws5, [28, 10, 12, 14, 16, 80])
    ws5.freeze_panes = "A3"

    # ── Sheet 6: Unused Triggers ───────────────────────────────────────────
    ws6 = wb.create_sheet("Unused Triggers")
    ws6.sheet_view.showGridLines = False
    _title(ws6, f"Unused Triggers  |  {len(unused_trigs)} of {len(triggers)} triggers unused", 4)
    for i, col in enumerate(["Trigger ID","Trigger Name","Type","Recommendation"], 1):
        _hdr(ws6, 2, i, col)
    for ri, tr in enumerate(unused_trigs, 3):
        ws6.row_dimensions[ri].height = 24
        bg = LT_AMB if ri%2==0 else LT_GRY
        n = tr.get("name","").lower()
        if any(kw in n for kw in ("cookie","consent","block","allow")):
            rec = "Remove if consent mechanism has changed. Confirm no custom JS references this trigger."
        elif any(kw in n for kw in ("old","backup","copy","deprecated","temp","test")):
            rec = "Remove — appears legacy or temporary."
        elif any(kw in n for kw in ("click","scroll","video","engagement")):
            rec = "Remove — interaction trigger with no tags attached."
        else:
            rec = "Remove in next cleanup sprint. Verify not referenced in custom JS variables first."
        _cell(ws6, ri, 1, tr["triggerId"],   bg=bg, sz=9, wrap=False)
        _cell(ws6, ri, 2, tr.get("name",""), bg=bg, sz=9)
        _cell(ws6, ri, 3, tr.get("type",""), bg=bg, sz=9, wrap=False)
        _cell(ws6, ri, 4, rec,               bg=bg, sz=9)
    _borders(ws6, 2, len(unused_trigs) + 2, 1, 4)
    _widths(ws6, [10, 54, 20, 72])
    ws6.freeze_panes = "A3"

    # Return bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════════

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700&display=swap');

.stFileUploader label,
div[data-testid="stFileUploader"] label,
div[data-testid="stFileUploaderDropzoneInstructions"] span {
    font-family: "Manrope", sans-serif !important;
}
div[data-testid="stPopover"] button {
    align-items: center; color: #8ec5ff; background: none; border: 0;
    display: inline-flex; gap: 0.25rem; justify-content: flex-start;
    padding: 0; font-size: 0.95rem; width: auto;
}
div[data-testid="stPopover"] button:hover { color: #8ec5ff; }
div[data-testid="stPopover"] button p { margin: 0; text-decoration: underline; }

/* Severity badge pills */
.badge {
    display: inline-block; padding: 2px 10px; border-radius: 12px;
    font-size: 0.75rem; font-weight: 600; letter-spacing: 0.04em;
    text-transform: uppercase; margin-right: 4px;
}
.badge-critical { background: #fadbd8; color: #c0392b; }
.badge-high     { background: #fdebd0; color: #d4881c; }
.badge-medium   { background: #d6eaf8; color: #1565c0; }
.badge-hygiene  { background: #d5f5e3; color: #1a6b3c; }

/* Finding cards */
.finding-card {
    border-left: 4px solid; border-radius: 4px; padding: 12px 16px;
    margin-bottom: 10px; background: rgba(255,255,255,0.03);
}
.finding-critical { border-color: #c0392b; }
.finding-high     { border-color: #d4881c; }
.finding-medium   { border-color: #1565c0; }
.finding-hygiene  { border-color: #1a6b3c; }

/* Metric cards */
.metric-row { display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 110px; border-radius: 8px; padding: 16px 12px 12px;
    text-align: center; border-top: 4px solid;
}
.metric-num  { font-size: 2.4rem; font-weight: 700; line-height: 1; margin: 0; }
.metric-lbl  { font-size: 0.72rem; font-weight: 600; letter-spacing: 0.08em;
               text-transform: uppercase; margin-top: 4px; }
.mc-critical { border-color:#c0392b; background:#fadbd8; color:#c0392b; }
.mc-high     { border-color:#d4881c; background:#fdebd0; color:#d4881c; }
.mc-medium   { border-color:#1565c0; background:#d6eaf8; color:#1565c0; }
.mc-hygiene  { border-color:#1a6b3c; background:#d5f5e3; color:#1a6b3c; }
</style>
"""


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    st.set_page_config(page_title="GTM Auto-Auditor", layout="wide")
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    st.title("GTM Auto-Auditor")
    st.markdown(
        "Upload a Google Tag Manager container export to audit your analytics "
        "implementation. Choose a tab below for the view you need."
    )

    help_text = (
        "In GTM, go to Admin > Export Container, select a workspace or version, "
        "and click Download to save a .json file."
    )
    if hasattr(st, "popover"):
        with st.popover("How do I get my GTM container?"):
            st.write(help_text)
    else:
        st.caption(help_text)

    st.divider()

    tab_audit, tab_inventory = st.tabs(["📋 Phase 1 Audit", "🔍 Analytics Inventory"])

    # ── SESSION STATE ──────────────────────────────────────────────────────
    for key, default in [
        ("audit_data", None), ("audit_sig", None),
        ("inventory_data", None), ("inventory_sig", None),
        ("show_full_inventory", False),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ══════════════════════════════════════════════════════════════════════
    # TAB 1: PHASE 1 AUDIT
    # ══════════════════════════════════════════════════════════════════════
    with tab_audit:
        st.subheader("Phase 1 Container Audit")
        st.markdown(
            "Upload a GTM container JSON export to receive a full structured audit: "
            "severity-tiered findings, a tag inventory, vendor summary, and a "
            "downloadable Excel workbook."
        )

        audit_file = st.file_uploader(
            "Upload GTM container JSON",
            type=["json"],
            accept_multiple_files=False,
            key="audit_uploader",
        )

        if audit_file is not None:
            try:
                file_bytes = audit_file.read()
                sig = hashlib.sha256(file_bytes).hexdigest()

                # Only reparse if the file has changed
                if st.session_state.audit_sig != sig:
                    st.session_state.audit_data = json.loads(file_bytes)
                    st.session_state.audit_sig  = sig

                container = _load_container_dict(st.session_state.audit_data)
                findings  = generate_findings(container)

                tags      = container["tags"]
                triggers  = container["triggers"]
                live_tags = [t for t in tags if not t.get("paused")]
                sev_counts = Counter(f["severity"] for f in findings)
                n_ap = len([t for t in live_tags if any(tid in ALL_PAGES_TRIGGERS for tid in t.get("firingTriggerId", []))])
                n_ua = len([t for t in live_tags if t["type"] == "ua"])
                n_html = len([t for t in live_tags if t["type"] == "html"])

                # ── Container stats bar ────────────────────────────────
                st.markdown("---")
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Total Tags",      len(tags))
                c2.metric("Triggers",        len(triggers))
                c3.metric("All-Pages Live",  n_ap)
                c4.metric("Live UA Tags",    n_ua,  delta=f"{'remove' if n_ua > 0 else 'none'}" if n_ua > 0 else None, delta_color="inverse")
                c5.metric("Custom HTML",     n_html)

                # ── Severity summary cards ─────────────────────────────
                st.markdown("#### Findings Summary")
                st.markdown(
                    f"""<div class="metric-row">
                      <div class="metric-card mc-critical">
                        <div class="metric-num">{sev_counts.get('Critical', 0)}</div>
                        <div class="metric-lbl">Critical</div>
                      </div>
                      <div class="metric-card mc-high">
                        <div class="metric-num">{sev_counts.get('High', 0)}</div>
                        <div class="metric-lbl">High</div>
                      </div>
                      <div class="metric-card mc-medium">
                        <div class="metric-num">{sev_counts.get('Medium', 0)}</div>
                        <div class="metric-lbl">Medium</div>
                      </div>
                      <div class="metric-card mc-hygiene">
                        <div class="metric-num">{sev_counts.get('Hygiene', 0)}</div>
                        <div class="metric-lbl">Hygiene</div>
                      </div>
                    </div>""",
                    unsafe_allow_html=True,
                )

                # ── Download button (generated immediately) ────────────
                xlsx_bytes = build_audit_workbook(container, findings)
                safe_id = re.sub(r"[^\w\-]", "_", container["container_id"])
                st.download_button(
                    label="⬇ Download Excel Workbook",
                    data=xlsx_bytes,
                    file_name=f"{safe_id}_audit.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

                st.markdown("---")

                # ── Findings by severity ───────────────────────────────
                st.markdown("#### Findings Detail")

                SEV_COLORS = {
                    "Critical": ("critical", "🔴"),
                    "High":     ("high",     "🟠"),
                    "Medium":   ("medium",   "🟡"),
                    "Hygiene":  ("hygiene",  "🟢"),
                }

                for sev in ["Critical", "High", "Medium", "Hygiene"]:
                    sev_findings = [f for f in findings if f["severity"] == sev]
                    if not sev_findings:
                        continue
                    cls, icon = SEV_COLORS[sev]
                    with st.expander(f"{icon} {sev} — {len(sev_findings)} finding{'s' if len(sev_findings) != 1 else ''}", expanded=(sev in ("Critical", "High"))):
                        for f in sev_findings:
                            st.markdown(
                                f"""<div class="finding-card finding-{cls}">
                                  <strong>{f['id']} · {f['title']}</strong><br>
                                  <span style="font-size:0.82rem;opacity:0.7;">{f['category']} · Owner: {f['owner']}</span>
                                </div>""",
                                unsafe_allow_html=True,
                            )
                            col_a, col_b = st.columns(2)
                            with col_a:
                                st.markdown("**Affected**")
                                st.caption(f["affected"])
                                st.markdown("**Why it matters**")
                                st.caption(f["why"])
                            with col_b:
                                st.markdown("**Recommended action**")
                                st.caption(f["action"])
                            st.markdown("---")

                # ── Quick-view tag table (filterable) ─────────────────
                st.markdown("#### Tag Inventory")
                trig_names = _build_trigger_map(triggers)
                tag_rows = []
                for t in tags:
                    is_orphan = not t.get("firingTriggerId") and not t.get("paused")
                    tag_rows.append({
                        "Tag ID":   t["tagId"],
                        "Name":     t["name"],
                        "Vendor":   _classify_vendor(t),
                        "Type":     t["type"],
                        "Status":   "Paused" if t.get("paused") else ("Orphaned" if is_orphan else "Live"),
                        "Scope":    _load_scope(t, trig_names),
                    })
                tag_df = pd.DataFrame(tag_rows)

                col_f1, col_f2 = st.columns(2)
                with col_f1:
                    vendor_opts = ["All"] + sorted(tag_df["Vendor"].unique().tolist())
                    sel_vendor  = st.selectbox("Filter by vendor", vendor_opts, key="audit_vendor_filter")
                with col_f2:
                    status_opts = ["All"] + sorted(tag_df["Status"].unique().tolist())
                    sel_status  = st.selectbox("Filter by status", status_opts, key="audit_status_filter")

                filtered = tag_df.copy()
                if sel_vendor != "All":
                    filtered = filtered[filtered["Vendor"] == sel_vendor]
                if sel_status != "All":
                    filtered = filtered[filtered["Status"] == sel_status]

                st.dataframe(filtered, use_container_width=True, hide_index=True)

            except json.JSONDecodeError:
                st.error("Unable to parse the uploaded file. Please ensure it is a valid JSON export from GTM.")
            except Exception as e:
                st.error(f"An error occurred while processing the file: {e}")
                st.exception(e)
        else:
            st.info("Upload a GTM container JSON file above to run the Phase 1 audit.")

    # ══════════════════════════════════════════════════════════════════════
    # TAB 2: ANALYTICS INVENTORY  (original feature, unchanged)
    # ══════════════════════════════════════════════════════════════════════
    with tab_inventory:
        st.subheader("Analytics Tag Inventory")
        st.markdown(
            "Upload a GTM container export to get a filterable table of GA4 and "
            "Universal Analytics tags with their trigger details. "
            "Optionally expand to the full tag inventory."
        )

        inv_file = st.file_uploader(
            "Upload GTM container JSON",
            type=["json"],
            accept_multiple_files=False,
            key="inventory_uploader",
        )

        if inv_file is not None:
            try:
                file_bytes = inv_file.read()
                sig = hashlib.sha256(file_bytes).hexdigest()
                if st.session_state.inventory_sig != sig:
                    st.session_state.inventory_data  = json.loads(file_bytes)
                    st.session_state.inventory_sig   = sig
                    st.session_state.show_full_inventory = False
                container_data = st.session_state.inventory_data

                df = parse_gtm_container(container_data, analytics_only=True)
                if df.empty:
                    st.warning(
                        "No Google Analytics tags were found in this container. "
                        "Please check that you exported the correct container."
                    )
                else:
                    tag_types    = ["All"] + sorted(df["Tag Type"].unique().tolist())
                    selected_type = st.selectbox("Filter by Tag Type", tag_types, key="inv_type_filter")
                    filtered_df  = df[df["Tag Type"] == selected_type] if selected_type != "All" else df
                    st.subheader("Tag Inventory")
                    st.dataframe(filtered_df, use_container_width=True)
                    st.download_button(
                        label="Download CSV",
                        data=filtered_df.to_csv(index=False),
                        file_name="gtm_analytics_inventory.csv",
                        mime="text/csv",
                    )

                st.divider()
                st.subheader("Full GTM Inventory")
                st.write("Generate a complete inventory of all tags in the container, not just analytics tags.")
                if st.button("Run Full GTM Inventory", use_container_width=True):
                    st.session_state.show_full_inventory = True

                if st.session_state.show_full_inventory and container_data is not None:
                    full_df = parse_gtm_container(container_data, analytics_only=False)
                    st.dataframe(full_df, use_container_width=True)
                    st.download_button(
                        label="Download Full GTM CSV",
                        data=full_df.to_csv(index=False),
                        file_name="gtm_full_inventory.csv",
                        mime="text/csv",
                    )
            except json.JSONDecodeError:
                st.error("Unable to parse the uploaded file. Please ensure it is a valid JSON export from GTM.")
            except Exception as e:
                st.error(f"An error occurred while processing the file: {e}")
        else:
            st.info("Upload a GTM container JSON file above to get started.")


if __name__ == "__main__":
    main()
